#Abre arquivo exel e coloca no destino ind1 ind2 data
import openpyxl 
from pathlib import Path

def escreve_casamentos(fonte, destino):
    arq_d = open(destino, "w")
    arq_d.close()
    arq_d = open(destino, "a")
    wb_obj = openpyxl.load_workbook(fonte) 
    
    # Read the active sheet:
    sheet = wb_obj['Casamento']

    for i in range(1, sheet.max_row):
        index_d = "A"
        index_h = "B"
        index_w = "G"
        
        index_d += str(i+1)
        index_h += str(i+1)
        index_w += str(i+1)
        str_ = str(sheet[index_h].value)
        str_ += ", "
        str_ += str(sheet[index_w].value)
        str_ += ", "
        str_c = str(sheet[index_d].value)
        if (str_c != "None"):
            str_c = str_c.split(" ")
            str_ += str_c[0]
        else:
            str_+= str_c
        if (i != sheet.max_row - 1):
            str_ += "\n"
        
        arq_d.write(str_)


escreve_casamentos('Fonte\\Sendon-Comp1.xlsx', "Fonte\\casamentos_rede_grande_c.txt")

